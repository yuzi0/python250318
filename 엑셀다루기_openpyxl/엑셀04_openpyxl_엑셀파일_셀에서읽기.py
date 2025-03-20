import  openpyxl  as  op  

#샘플로 제공된 파일을 오픈한다. 
wb = op.load_workbook("result2.xlsx") 
# #마지막으로 활성화된 시트를 가져온다
# ws = wb.active
# 명시적으로 시트명 지정
ws = wb["직원명부"]

#방법 1 : Sheet의 Cell 속성 사용하기
data1 = ws.cell(row=1, column=2).value

#방법 2 : 엑셀 인덱스(Range) 사용하기
data2 = ws["B1"].value

#위 결과 출력해보기
print("cell(1,2) : ", data1)
print('Range("B1"):', data2)

#범위를 지정한 경우 
rng = ws["A1:C3"] 

for  rng_data  in  rng: 
    for  cell_data  in  rng_data: 
        print(cell_data.value) 

